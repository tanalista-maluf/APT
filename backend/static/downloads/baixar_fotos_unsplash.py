# -*- coding: utf-8 -*-
"""
Baixa as 375 fotos do Unsplash (via API oficial) para as pastas do Round 2
e grava o nome de cada arquivo baixado de volta no Excel.

COMO USAR
1. Instale as dependências (uma vez só):
   pip3 install requests openpyxl --break-system-packages

2. Preencha as 3 variáveis logo abaixo (UNSPLASH_ACCESS_KEY, XLSX_PATH, ROUND2_PATH).

3. Rode no Terminal:
   python3 baixar_fotos_unsplash.py

O script é seguro para rodar mais de uma vez: se um arquivo já existe na pasta
de destino, ele pula o download e só garante que o nome está gravado no Excel.
"""

import os
import re
import time
import unicodedata
import requests
import openpyxl

# ---------- CONFIGURAÇÃO (edite estas 3 linhas) ----------
UNSPLASH_ACCESS_KEY = "COLOQUE_SUA_ACCESS_KEY_AQUI"
XLSX_PATH = os.path.expanduser("~/Desktop/Round 2/conteudo_expedicoes_instagram.xlsx")
ROUND2_PATH = os.path.expanduser("~/Desktop/Round 2")
# -----------------------------------------------------------

# Fotos que o Unsplash não encontrou ficam registradas aqui (célula do Excel
# continua em branco, o script para de tentar de novo pra não gastar cota
# à toa) — abra esse arquivo pra ver o que falta resolver manualmente.
RELATORIO_PENDENTES_PATH = os.path.join(ROUND2_PATH, "fotos_sem_resultado.txt")

TEMA_PARA_SUBPASTA = {
    "5 Cidades": "1.Cidades",
    "5 Highlights": "2.Highlights",
    "5 Pratos e Bebidas Típicas": "3.Pratos e Bebidas",
    "5 Locais Instagramáveis": "4.Locais",
    "5 Dicas Práticas": "5.Dicas",
}
TEMAS_ORDEM = list(TEMA_PARA_SUBPASTA.keys())

# mesmas queries usadas para gerar os links de busca do Unsplash
QUERIES = {
"Al-Andalus": {
    "5 Cidades": ["Seville Spain cityscape","Cordoba Spain old town","Ronda Spain bridge","Granada Spain Alhambra hill","Malaga Spain castle"],
    "5 Highlights": ["Giralda Seville tower","Mezquita Cordoba interior","Medina Azahara ruins","Alhambra Nasrid Palace","Albaicin Granada sunset"],
    "5 Pratos e Bebidas Típicas": ["salmorejo Spanish soup","jamon iberico ham","pescaito frito Andalusia","Jerez sherry wine","Moroccan mint tea Granada"],
    "5 Locais Instagramáveis": ["Court of the Lions Alhambra","Mirador San Nicolas view","Mezquita Cordoba columns","Puente Nuevo Ronda","Albaicin narrow streets"],
    "5 Dicas Práticas": ["Alhambra Granada tourists","Andalusia cobblestone streets walking","Mirador de San Nicolas sunset","Andalusia summer heat","Spain tapas bar"],
},
"Caminhos de Cervantes": {
    "5 Cidades": ["Madrid Spain skyline","Alcala de Henares university","Toledo Spain old town","Consuegra windmills","Campo de Criptana windmills"],
    "5 Highlights": ["Cervantes birthplace museum","Toledo Cathedral interior","El Greco painting museum","Consuegra windmills La Mancha","Campo de Criptana village"],
    "5 Pratos e Bebidas Típicas": ["cochinillo roast suckling pig","Manchego cheese Spain","La Mancha vineyard wine","migas Spanish dish","torrijas Spanish dessert"],
    "5 Locais Instagramáveis": ["windmills sunset La Mancha","Toledo narrow alley","Sinagoga del Transito Toledo","Campo de Criptana windmills","Plaza Mayor Madrid"],
    "5 Dicas Práticas": ["La Mancha windmills sunset","Toledo walking tourists","museum Toledo entrance","Manchego cheese shop","Toledo steep street"],
},
"Escandinávia Boreal": {
    "5 Cidades": ["Reykjavik Iceland city","Vik Iceland black beach","Tromso Norway city","Kiruna Sweden Ice Hotel","Inari Finland Lapland"],
    "5 Highlights": ["Blue Lagoon Iceland","Jokulsarlon glacier lagoon","aurora borealis Tromso","Ice Hotel Kiruna room","husky sled Lapland"],
    "5 Pratos e Bebidas Típicas": ["skyr Icelandic yogurt","fresh fish Reykjavik market","reindeer dish Lapland","hot tea winter Arctic","Nordic lamb dish"],
    "5 Locais Instagramáveis": ["Jokulsarlon icebergs","Diamond Beach Iceland","Arctic Cathedral Tromso","Ice Hotel Kiruna interior","aurora borealis sky Finland"],
    "5 Dicas Práticas": ["Arctic winter clothing layers","aurora borealis night sky","camera tripod northern lights","Ice Hotel sleeping bag","northern lights photography"],
},
"A Sombra da Guerra": {
    "5 Cidades": ["Paris France cityscape","Caen Normandy France","Munich Germany city","Nuremberg Germany old town","Berlin Germany skyline"],
    "5 Highlights": ["Omaha Beach Normandy","Colleville American Cemetery","Dachau memorial","Nuremberg trials courtroom","Holocaust Memorial Berlin"],
    "5 Pratos e Bebidas Típicas": ["Normandy crepes France","Bavarian beer Munich","currywurst Berlin street food","Bavaria wine region","German bakery bread"],
    "5 Locais Instagramáveis": ["Colleville sur Mer cemetery crosses","East Side Gallery Berlin Wall","Reichstag glass dome Berlin","Brandenburg Gate Berlin","Pointe du Hoc cliffs"],
    "5 Dicas Práticas": ["Normandy beach memorial visitors","D-Day sites walking tour","Normandy beach uneven terrain","World War II history book","outdoor memorial site water"],
},
"Route des Vins": {
    "5 Cidades": ["Reims France cathedral","Strasbourg France old town","Dijon France city","Bordeaux France city","Cognac France town"],
    "5 Highlights": ["Reims Cathedral France","Avenue de Champagne Epernay","Hospices de Beaune","Saint Emilion village","Chateau Yquem Sauternes"],
    "5 Pratos e Bebidas Típicas": ["Champagne bottles France","Alsace Riesling wine","Dijon mustard","Bordeaux wine bottle","Cognac barrels cellar"],
    "5 Locais Instagramáveis": ["Champagne vineyards France","Colmar Alsace colorful houses","Saint Emilion medieval village","Sauternes vineyard chateau","Cognac wine cellar barrels"],
    "5 Dicas Práticas": ["wine tasting vineyard France","wine tour small group","vineyard boots walking","wine tasting glass water","wine cellar shopping"],
},
"Terra do Gelo e do Fogo": {
    "5 Cidades": ["Reykjavik Iceland","Thingvellir national park Iceland","Vik Iceland black sand beach","Katla glacier Iceland","Jokulsarlon Iceland lagoon"],
    "5 Highlights": ["Golden Circle Iceland","Reynisfjara black beach","Katla ice cave blue","Diamond Beach Iceland icebergs","Strokkur geyser eruption"],
    "5 Pratos e Bebidas Típicas": ["skyr Icelandic yogurt","fresh fish Iceland","Icelandic lamb dish","hot chocolate winter","Icelandic coffee"],
    "5 Locais Instagramáveis": ["Diamond Beach Iceland ice","blue ice cave Iceland","Reynisfjara basalt columns","Seljalandsfoss waterfall behind","Gullfoss waterfall Iceland"],
    "5 Dicas Práticas": ["ice cave crampons helmet","Reynisfjara waves warning","Golden Circle tourists morning","Iceland rain gear","camera battery cold weather"],
},
"Pop Nórdico": {
    "5 Cidades": ["Stockholm Sweden city","Gamla Stan Stockholm","Halmstad Sweden town","Gothenburg Sweden city","Stockholm waterfront sunset"],
    "5 Highlights": ["ABBA Museum Stockholm","Berns Salonger Stockholm","Halmstad Sweden street","Bengans record store Gothenburg","recording studio Sweden"],
    "5 Pratos e Bebidas Típicas": ["Swedish meatballs kottbullar","fika coffee cinnamon bun","pickled herring Sweden","gravlax salmon Sweden","Swedish chocolate"],
    "5 Locais Instagramáveis": ["ABBA Museum entrance colorful","Gamla Stan colorful houses","Malaren lake Stockholm sunset","Halmstad old town","record store vinyl shop"],
    "5 Dicas Práticas": ["ABBA Museum tourists","fika cafe Sweden","train Sweden travel","Sweden road trip playlist","lake sunset photography Stockholm"],
},
"Reinos Dourados": {
    "5 Cidades": ["Bangkok Thailand skyline","Chiang Mai Thailand temple","Krabi Thailand coast","Phi Phi islands Thailand","Siem Reap Cambodia"],
    "5 Highlights": ["Grand Palace Bangkok","elephant sanctuary Chiang Mai","Phi Phi islands boat","Angkor Wat sunrise","Ta Prohm temple roots"],
    "5 Pratos e Bebidas Típicas": ["pad thai dish","Thai curry dish","tropical fruit market Thailand","Cambodian fish amok","Khmer coffee Cambodia"],
    "5 Locais Instagramáveis": ["Wat Arun sunset Bangkok","Railay beach Thailand cliffs","Angkor Wat reflection lake","Ta Prohm tree roots temple","Tonle Sap floating village"],
    "5 Dicas Práticas": ["Angkor Wat sunrise crowd","tropical sun protection travel","temple modest dress Asia","Cambodia visa document","tropical humidity hydration travel"],
},
"Espanha Medieval": {
    "5 Cidades": ["Girona Spain old town","Cardona Spain castle","Albarracin Spain red village","Toledo Spain skyline","Segovia Spain aqueduct"],
    "5 Highlights": ["Girona Jewish quarter","Cardona salt castle","Albarracin red village Spain","Toledo three cultures monument","Avila city walls Spain"],
    "5 Pratos e Bebidas Típicas": ["pa amb tomaquet Catalan","Manchego cheese Spain","roast lamb Castile","Ribera del Duero wine","churros con chocolate"],
    "5 Locais Instagramáveis": ["Albarracin village wall view","Segovia Roman aqueduct","Avila city walls sunrise","Toledo skyline viewpoint","Montserrat monastery Spain"],
    "5 Dicas Práticas": ["medieval cobblestone street walking","Toledo viewpoint sunset","castle entrance ticket line","Spain spring coat evening","churros breakfast Spain"],
},
"Celtic Myths e Whisky": {
    "5 Cidades": ["Edinburgh Scotland skyline","Speyside Scotland distillery","Isle of Skye Scotland","Islay Scotland coast","Glasgow Scotland city"],
    "5 Highlights": ["Edinburgh Castle Scotland","Speyside whisky distillery","Neist Point Skye","Talisker distillery Skye","Islay whisky distillery"],
    "5 Pratos e Bebidas Típicas": ["Scotch whisky single malt","haggis Scottish dish","smoked salmon Scotland","shortbread Scottish","craft beer Scotland pub"],
    "5 Locais Instagramáveis": ["Neist Point lighthouse Skye","Old Man of Storr Skye","Eilean Donan Castle Scotland","Islay coastline cliffs","Royal Mile Edinburgh"],
    "5 Dicas Práticas": ["whisky tasting distillery tour","Scotland windbreaker jacket hiking","ferry Islay weather","whisky tasting water glass","hiking boots Skye trail"],
},
"Vinhos e Reinos": {
    "5 Cidades": ["Rioja Spain vineyard","Barcelona Spain skyline","Porto Portugal riverside","Douro Valley Portugal","Lisbon Portugal skyline"],
    "5 Highlights": ["Rioja bodega wine cellar","Priorat vineyard Spain","Porto wine cellars Gaia","Douro Valley terraces","Sintra Portugal palace"],
    "5 Pratos e Bebidas Típicas": ["Rioja red wine Spain","Port wine Porto","pastel de nata Lisbon","bacalhau codfish Portugal","Alentejo cheese Portugal"],
    "5 Locais Instagramáveis": ["Douro Valley terraced vineyards","Porto Ribeira colorful houses","Sintra colorful palace","Priorat vineyard slate soil","Belem Tower Lisbon"],
    "5 Dicas Práticas": ["Douro river boat tour","Port wine cellar tasting Gaia","Douro Valley sun hat","Sintra morning crowd","Lisbon tram travel"],
},
"Rota da Transilvânia": {
    "5 Cidades": ["Bucharest Romania city","Brasov Romania old town","Sighisoara Romania medieval","Sibiu Romania houses","Sofia Bulgaria city"],
    "5 Highlights": ["Peles Castle Romania","Bran Castle Romania","Sighisoara clock tower","Viscri fortified church Romania","Rila Monastery Bulgaria"],
    "5 Pratos e Bebidas Típicas": ["sarmale Romanian dish","Bulgarian beer","mici Romanian grilled sausages","Carpathian goat cheese","Balkan local wine"],
    "5 Locais Instagramáveis": ["Bran Castle Romania view","Sighisoara clock tower colorful","Sibiu eyes houses facade","Alexander Nevsky Cathedral Sofia","Viscri village Romania"],
    "5 Dicas Práticas": ["Carpathian mountains cold night","Peles Castle interior photo","medieval cobblestone Sighisoara","Romanian street food market","border crossing passport travel"],
},
"La Dolce Vite": {
    "5 Cidades": ["Rome Italy skyline","Orvieto Italy hilltop town","Florence Italy skyline","Siena Italy old town","Montalcino Italy vineyard"],
    "5 Highlights": ["Roman Forum Colosseum","Sistine Chapel Vatican","Uffizi Gallery Florence","Piazza del Campo Siena","Brunello di Montalcino winery"],
    "5 Pratos e Bebidas Típicas": ["Roman pasta carbonara","Brunello di Montalcino wine","Florence gelato","Vino Nobile Montepulciano wine","Tuscan olive oil"],
    "5 Locais Instagramáveis": ["Val d'Orcia Tuscany cypress","San Gimignano towers Italy","Florence Duomo skyline","Roman Forum sunset","Montalcino vineyard hills"],
    "5 Dicas Práticas": ["Uffizi Gallery ticket line","Val d'Orcia sunset photography","Rome morning walking tourists","gelato shop Florence","cobblestone street Tuscany"],
},
"Abbeys e Ales": {
    "5 Cidades": ["Brussels Belgium Grand Place","Ghent Belgium canal","Bruges Belgium canal","Westmalle Belgium abbey","Chimay Belgium countryside"],
    "5 Highlights": ["Grand Place Brussels","Ghent Altarpiece painting","Bruges canal boat","Westmalle Abbey Belgium","Orval Abbey ruins"],
    "5 Pratos e Bebidas Típicas": ["Trappist beer Belgium","Belgian waffles","Belgian chocolate shop","mussels and fries Belgium","abbey cheese Belgium"],
    "5 Locais Instagramáveis": ["Grand Place Brussels night","Bruges canal houses","Orval Abbey ruins Belgium","Ghent colorful facades","Trappist beer glass abbey"],
    "5 Dicas Práticas": ["Trappist beer tasting","abbey tour Belgium visitors","Bruges canal boat tour","Belgian waffle street vendor","Belgium rainy weather coat"],
},
"Fab Four Tour": {
    "5 Cidades": ["Liverpool England city","London England skyline","Stratford upon Avon England","Abbey Road London street","Piccadilly Circus London"],
    "5 Highlights": ["Cavern Club Liverpool","Abbey Road Studios London","Carnaby Street London","Albert Dock Liverpool","Oxford University England"],
    "5 Pratos e Bebidas Típicas": ["fish and chips England","English afternoon tea","British ale pub","Sunday roast England","scones clotted cream"],
    "5 Locais Instagramáveis": ["Abbey Road crossing London","Cavern Club Liverpool interior","Carnaby Street colorful London","Oxford gothic architecture","Albert Dock sunset Liverpool"],
    "5 Dicas Práticas": ["Abbey Road crossing traffic","Cavern Club live music night","London umbrella rain","Oxford walking tour","afternoon tea reservation London"],
},
}

# colunas de imagem no Excel: item 1..5 -> colunas D,G,J,M,P (4,7,10,13,16)
IMG_COLS = [4, 7, 10, 13, 16]


def normalizar(txt):
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode()
    txt = txt.lower()
    txt = re.sub(r"[^a-z0-9]+", "", txt)
    return txt


def mapear_pastas_expedicao(round2_path, nomes_planilhas):
    """Casa o nome de cada aba do Excel com a subpasta correspondente dentro de Round 2."""
    pastas_existentes = [d for d in os.listdir(round2_path) if os.path.isdir(os.path.join(round2_path, d))]
    norm_to_real = {normalizar(p): p for p in pastas_existentes}
    mapa = {}
    faltando = []
    for nome in nomes_planilhas:
        alvo = norm_to_real.get(normalizar(nome))
        if alvo:
            mapa[nome] = alvo
        else:
            faltando.append(nome)
    return mapa, faltando


def carregar_pendentes(caminho):
    """Le o relatorio de fotos sem resultado e devolve:
    - chaves: set de "sheet||tema||item" ja tentados sem sucesso (pra pular)
    - detalhes: dict chave -> texto legivel (pra reimprimir no relatorio)
    """
    chaves, detalhes = set(), {}
    if not os.path.exists(caminho):
        return chaves, detalhes
    with open(caminho, "r", encoding="utf-8") as f:
        for linha in f:
            linha = linha.strip()
            if not linha or linha.startswith("#"):
                continue
            partes = linha.split(" | ")
            if len(partes) < 4:
                continue
            sheet, tema, item_txt, query = partes[0], partes[1], partes[2], " | ".join(partes[3:])
            chave = f"{sheet}||{tema}||{item_txt}"
            chaves.add(chave)
            detalhes[chave] = linha
    return chaves, detalhes


def salvar_pendentes(caminho, detalhes):
    linhas = sorted(detalhes.values())
    with open(caminho, "w", encoding="utf-8") as f:
        f.write("# Fotos que o Unsplash não encontrou (célula do Excel fica em branco).\n")
        f.write("# Resolva manualmente: busque no unsplash.com, baixe, salve na pasta do\n")
        f.write("# tema com o nome do arquivo, e cole esse nome na célula correspondente\n")
        f.write("# do Excel. Depois pode apagar a linha correspondente deste arquivo (ou\n")
        f.write("# deixar - ele só é usado pra não tentar de novo automaticamente).\n\n")
        for linha in linhas:
            f.write(linha + "\n")


def remover_pendente_resolvido(detalhes, sheet_name, tema, item_num):
    """Se o usuário já preencheu manualmente essa célula no Excel, tira da lista de pendentes."""
    chave = f"{sheet_name}||{tema}||item {item_num}"
    detalhes.pop(chave, None)


class RateLimitError(Exception):
    """Levantado quando a API do Unsplash sinaliza que o limite por hora acabou."""
    pass


def buscar_foto(query, headers):
    """Retorna (foto, cota_zerada). 'foto' é None se não houver resultado."""
    resp = requests.get(
        "https://api.unsplash.com/search/photos",
        headers=headers,
        params={"query": query, "per_page": 1},
        timeout=20,
    )

    if resp.status_code == 403:
        raise RateLimitError("A API recusou a chamada (limite de requisições excedido).")

    resp.raise_for_status()
    data = resp.json()

    restante = resp.headers.get("X-Ratelimit-Remaining")
    cota_zerada = restante is not None and int(restante) <= 0

    foto = data["results"][0] if data.get("results") else None
    return foto, cota_zerada


def montar_nome_arquivo(foto):
    username = foto.get("user", {}).get("username", "unsplash")
    photo_id = foto.get("id", "")
    return f"{username}-{photo_id}-unsplash.jpg"


def baixar_arquivo(url, destino):
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    with open(destino, "wb") as f:
        f.write(resp.content)


def main():
    if UNSPLASH_ACCESS_KEY == "COLOQUE_SUA_ACCESS_KEY_AQUI":
        print("Edite o arquivo e cole sua Access Key do Unsplash antes de rodar.")
        return

    headers = {"Authorization": f"Client-ID {UNSPLASH_ACCESS_KEY}"}

    wb = openpyxl.load_workbook(XLSX_PATH)
    nomes_planilhas = [n for n in wb.sheetnames if n in QUERIES]

    mapa_pastas, faltando = mapear_pastas_expedicao(ROUND2_PATH, nomes_planilhas)
    if faltando:
        print("ATENÇÃO — não encontrei subpasta em Round 2 para estas abas (pulando):")
        for f in faltando:
            print("  -", f)

    pendentes_chaves, pendentes_detalhes = carregar_pendentes(RELATORIO_PENDENTES_PATH)

    total = 0
    baixados = 0
    pulados = 0
    erros = 0
    ja_marcados = 0
    parou_por_limite = False

    for sheet_name in nomes_planilhas:
        if parou_por_limite:
            break
        if sheet_name not in mapa_pastas:
            continue
        pasta_expedicao = os.path.join(ROUND2_PATH, mapa_pastas[sheet_name])
        ws = wb[sheet_name]
        q_map = QUERIES[sheet_name]

        for r, tema in enumerate(TEMAS_ORDEM, start=2):
            if parou_por_limite:
                break
            subpasta_nome = TEMA_PARA_SUBPASTA[tema]
            pasta_tema = os.path.join(pasta_expedicao, subpasta_nome)
            if not os.path.isdir(pasta_tema):
                print(f"[AVISO] Subpasta não encontrada, pulando: {pasta_tema}")
                continue

            for i, query in enumerate(q_map[tema]):
                col = IMG_COLS[i]
                cell = ws.cell(row=r, column=col)

                # já tem nome de arquivo gravado e o arquivo existe -> nada a fazer
                if isinstance(cell.value, str) and cell.value.endswith((".jpg", ".jpeg", ".png")):
                    if os.path.exists(os.path.join(pasta_tema, cell.value)):
                        pulados += 1
                        total += 1
                        # se o usuário resolveu manualmente, tira da lista de pendentes
                        remover_pendente_resolvido(pendentes_detalhes, sheet_name, tema, i + 1)
                        continue

                # já tentamos essa foto antes e não teve resultado -> não gasta
                # cota de novo, só reaparece no relatório final
                chave = f"{sheet_name}||{tema}||item {i+1}"
                if chave in pendentes_chaves:
                    ja_marcados += 1
                    total += 1
                    continue

                total += 1

                try:
                    foto, cota_zerada = buscar_foto(query, headers)

                    if not foto:
                        print(f"[SEM RESULTADO] {sheet_name} / {tema} / item {i+1} ({query})")
                        pendentes_detalhes[chave] = f"{sheet_name} | {tema} | item {i+1} | {query}"
                        erros += 1
                    else:
                        nome_arquivo = montar_nome_arquivo(foto)
                        destino = os.path.join(pasta_tema, nome_arquivo)

                        if not os.path.exists(destino):
                            baixar_arquivo(foto["urls"]["raw"], destino)
                            try:
                                requests.get(foto["links"]["download_location"], headers=headers, timeout=10)
                            except requests.RequestException:
                                pass
                            baixados += 1
                        else:
                            pulados += 1

                        cell.value = nome_arquivo

                    if cota_zerada:
                        print("\n[LIMITE] Cota de 50 requisições/hora esgotada nesta chamada.")
                        parou_por_limite = True
                        break

                    time.sleep(1.3)

                except RateLimitError as e:
                    print(f"\n[LIMITE] {e}")
                    parou_por_limite = True
                    break

                except Exception as e:
                    erros += 1
                    print(f"[ERRO] {sheet_name} / {tema} / item {i+1} ({query}): {e}")

        wb.save(XLSX_PATH)  # salva incrementalmente a cada expedição

    wb.save(XLSX_PATH)
    salvar_pendentes(RELATORIO_PENDENTES_PATH, pendentes_detalhes)

    print("\nConcluído esta rodada." if not parou_por_limite else "\nPausado por limite de requisições.")
    print(f"Total de itens verificados: {total}")
    print(f"Baixados agora: {baixados} | Já existiam: {pulados} | Sem resultado (novos): {erros} | Já marcados como sem resultado: {ja_marcados}")
    if parou_por_limite:
        print("\nA cota de 50 requisições/hora do Unsplash acabou.")
        print("Espere a próxima hora e rode o script de novo — ele retoma automaticamente")
        print("de onde parou, sem baixar de novo o que já foi salvo.")

    if pendentes_detalhes:
        print(f"\n{'='*60}")
        print(f"FOTOS SEM RESULTADO ({len(pendentes_detalhes)}) — precisam de ação manual:")
        print(f"{'='*60}")
        for linha in sorted(pendentes_detalhes.values()):
            print(f"  • {linha}")
        print(f"\nLista completa salva em: {RELATORIO_PENDENTES_PATH}")
        print("Pra resolver: busque a foto no unsplash.com, salve na pasta do tema,")
        print("e cole o nome do arquivo na célula correspondente do Excel.")


if __name__ == "__main__":
    main()
